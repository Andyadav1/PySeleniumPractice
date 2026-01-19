import openpyxl

def UpdateExcelData(FilePath,FruitName,ColName,NewValue):
    book = openpyxl.load_workbook(FilePath)
    sheet = book.active
    cords = {}
    for i in range(1,sheet.max_column+1):
        if sheet.cell(1,i).value == ColName:
            cords["col"] = i

    for i in range(1,sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            if sheet.cell(i,j).value == FruitName:
                cords["row"] = i

    sheet.cell(row= cords["row"],column= cords["col"]).value = NewValue
    book.save(FilePath)


